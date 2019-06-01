# excel file 읽기
import openpyxl as e

exce = e.load_workbook("D:/python/read/score2.xlsx")
aws = exce.active
for i in aws.rows :
    index = i[0].row
    name = i[0].value
    kor = i[1].value
    eng = i[2].value
    math = i[3].value

    total = kor + eng + math
    avg = total/3
    print("{} {} {} {} {} {} {}".format(index,name,kor,eng,math, total,avg))

    aws.cell(row=index, column=6).value = total
    aws.cell(row=index, column=7).value = avg



exce.save("D:/python/read/result.xlsx")
exce.close()